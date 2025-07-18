import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def prepare_data(reports_df: pd.DataFrame, freequest_df: pd.DataFrame) -> pd.DataFrame:
    """報告と対応するクエストデータをマージする

    Args:
        reports_df (pd.DataFrame): DBから取得した報告
        freequest_df (pd.DataFrame): フリークエストデータ

    Returns:
        pd.DataFrame: マージされたデータ
    """
    # "category"が"Error"の行を除外
    reports_df = reports_df[reports_df["category"] != "Error"]

    reports_df_pivot = reports_df.pivot_table(
        index=["id", "runs", "note", "war_name", "quest_name", "url", "timestamp"],
        columns="object_name",
        values="num",
        fill_value=None,
    ).reset_index()

    # freequest_dfの"quest_name"を"counter_name"の値で置き換え
    freequest_df["quest_name"] = freequest_df["counter_name"]
    merged_df = freequest_df.merge(
        reports_df_pivot, on=["war_name", "quest_name"], how="outer"
    )
    return merged_df


def create_output_df(group: pd.DataFrame, item_columns: np.ndarray) -> pd.DataFrame:
    """あるフリクエ用にExcelの統計シート出力用にデータを整形する

    Args:
        group (pd.DataFrame): そのフリクエの全入力データ
        item_columns (np.ndarray): そのフリクエでドロップするアイテム

    Returns:
        pd.DataFrame: 整形されたデータ
    """
    # ソート操作を追加。timestamp列に基づいて昇順にソート
    group = group.sort_values(by="timestamp")

    columns_after_item13 = group.columns.tolist()[
        group.columns.tolist().index("item19") + 1
    ]
    output_columns = ["url", "timestamp", "runs"] + list(item_columns)

    if group[columns_after_item13].isnull().all().all():
        output_df = pd.DataFrame(columns=output_columns)
    else:
        output_df = group.drop(["id", "category", "war_name"], axis=1).reset_index(
            drop=True
        )

    for col in output_columns:
        if col not in output_df.columns:
            output_df[col] = ""

    output_df = output_df[output_columns[2:] + ["url", "timestamp"]]
    output_df = output_df.rename(
        columns={"url": "ソース", "timestamp": "メモ", "runs": "周回数"}
    )
    output_df = output_df.T
    output_df.reset_index(inplace=True)
    output_df.columns = pd.Index(["No."] + list(output_df.columns[1:]))
    return output_df


def write_to_sheet(
    ws: Worksheet,
    output_df: pd.DataFrame,
    war_name: str,
    quest_name: str,
    previous_war_name: str,
) -> str:
    """あるクエストの全データをまとめて特定の統計シートに出力する

    Args:
        ws (Worksheet): 出力するシート
        output_df (pd.DataFrame): quest_nameの全報告データ
        war_name (str): いわゆる特異点名、修練場の場合は曜日
        quest_name (str): クエスト名
        previous_war_name (str): このクエストの前に出力した特異点名

    Returns:
        str: _description_
    """
    if war_name != previous_war_name:
        ws.append([war_name])

    ws.append([None, quest_name])
    ws.append([None] + ["No."])

    for data_row in output_df.values:
        ws.append([None] + data_row.tolist()[:1] + [None, None] + data_row.tolist()[1:])
        if data_row.tolist()[0] == "メモ":
            for cell in ws[ws.max_row][4:]:
                cell.number_format = "mm/dd"
        elif data_row.tolist()[0] == "ソース":
            for cell in ws[ws.max_row][4:]:
                if cell.value is not None:
                    cell.hyperlink = str(cell.value)  # type: ignore

    ws.append([])
    ws.append([])
    return war_name  # war_nameを更新


def aggregate_items_by_object(df: pd.DataFrame) -> pd.DataFrame:
    """
    DataFrameを入力として受け取り、特定のカラムとobject_nameでグループ化し、
    numとstackの乗算の合計を計算して集計する関数。

    Args:
        df: 入力Pandas DataFrame。以下のカラムを含み、numとstackがint型であることを期待します。
            id, owner, name, twitter_id, twitter_name, twitter_username,
            report_type, war_name, quest_type, quest_name, timestamp,
            runs, note, object_name, num, stack, category, url

    Returns:
        集計済みのPandas DataFrame。object_nameごとにnum * stackの合計が
        計算され、stackは1に設定されます。元のDataFrameと同じカラム構成と順序になります。
    """
    # 入力DataFrameのコピーを作成し、元のDataFrameを変更しないようにする
    df_processed = df.copy()

    # numとstackの乗算結果を計算する一時カラムを作成
    # numとstackがint型であることを前提としていますが、乗算結果や合計は大きくなる可能性があるので、
    # Pandasに適切な型を推論させます。
    df_processed["_num_stack_product"] = df_processed["num"] * df_processed["stack"]

    # グループ化に使用するカラムリストを定義
    # object_nameもグループ化キーに含めます
    group_cols = [
        "id",
        "owner",
        "name",
        "twitter_id",
        "twitter_name",
        "twitter_username",
        "report_type",
        "war_name",
        "quest_type",
        "quest_name",
        "timestamp",
        "runs",
        "note",
        "category",
        "url",
        "object_name",
    ]

    # 指定されたカラムとobject_nameでグループ化し、_num_stack_productの合計を計算
    # aggメソッドを使用して集計し、結果のカラム名を'num'とする
    aggregated_df = (
        df_processed.groupby(group_cols)
        .agg(num=("_num_stack_product", "sum"))
        .reset_index()
    )  # グループ化キーをカラムに戻す

    # 集計後のDataFrameにstackカラムを追加し、すべての値を1に設定
    aggregated_df["stack"] = 1

    # 元のDataFrameの順序に合わせてカラムを並べ替える
    # 必要なカラムリストを定義
    output_cols = [
        "id",
        "owner",
        "name",
        "twitter_id",
        "twitter_name",
        "twitter_username",
        "report_type",
        "war_name",
        "quest_type",
        "quest_name",
        "timestamp",
        "runs",
        "note",
        "object_name",
        "num",
        "stack",
        "category",
        "url",
    ]

    # 並べ替えたDataFrameを作成
    # aggの結果のカラム順序は group_cols + ['num', 'stack'] のようになっているため、
    # output_colsリストを使って明示的にカラムを選択し、順序を調整します。
    result_df = aggregated_df[output_cols]

    return result_df


def create_statics(wb: Workbook, reports_df: pd.DataFrame, freequest_df: pd.DataFrame):
    """統計シートを出力する

    Args:
        wb (Workbook): 出力するワークブック
        reports_df (pd.DataFrame): 報告データ
        freequest_df (pd.DataFrame): フリークエストに関するデータ
    """
    new_report_df = aggregate_items_by_object(reports_df)
    merged_df = prepare_data(new_report_df, freequest_df)

    order = [
        "修練場",
        "フリクエ1部",
        "フリクエ1.5部",
        "フリクエ2部",
        "奏章",
        "冠位戴冠戦",
    ]
    # カテゴリごとに処理
    for category_name in order:
        group = freequest_df[freequest_df["category"] == category_name]
        ws = wb.create_sheet(title=f"統計【{category_name}】")  # 新しいシートを作成

        # 前回のwar_nameを記憶する変数
        previous_war_name = ""
        for _, row in group.iterrows():
            quest_name = row["counter_name"]
            war_name = row["war_name"]

            group = merged_df[merged_df["quest_name"] == quest_name]

            item_columns = freequest_df.loc[
                freequest_df["quest_name"] == quest_name,
                [
                    "item1",
                    "item2",
                    "item3",
                    "item4",
                    "item5",
                    "item6",
                    "item7",
                    "item8",
                    "item9",
                    "item10",
                    "item11",
                    "item12",
                    "item13",
                    "item14",
                    "item15",
                    "item16",
                    "item17",
                    "item18",
                    "item19",
                ],
            ].values.ravel()
            item_columns = item_columns[~pd.isnull(item_columns)]
            output_df = create_output_df(group, item_columns)
            previous_war_name = write_to_sheet(
                ws, output_df, war_name, quest_name, previous_war_name
            )


def append_rows_to_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """Worksheetに各行を追加する

    Args:
        ws (Worksheet): 出力するワークブックシート
        df (pd.DataFrame): 使用するデータフレーム
    """
    # ヘッダーを追加
    headers = [
        "id",
        "timestamp",
        "owner",
        "name",
        "twitter_id",
        "twitter_name",
        "twitter_username",
        "note",
        "url",
        "war_name",
        "quest_name",
        "runs",
    ]
    ws.append(headers)

    if df.empty:
        return
    id_group_df = df.groupby("id").agg(list)
    # タイムスタンプで降順再ソート
    id_group_df = id_group_df.sort_values(by="timestamp", ascending=False)

    for idx, row in id_group_df.iterrows():
        new_row = []
        for object_name, num, stack in zip(
            row["object_name"], row["num"], row["stack"]
        ):
            if stack == 1:
                new_row.extend([object_name, num])
            elif stack > 1:
                if (
                    object_name == "QP"
                    or object_name.endswith("ポイント")
                    or object_name.endswith("P")
                ):
                    new_row.extend([f"{object_name}(+{stack})", num])
                else:
                    new_row.extend([f"{object_name}(x{stack})", num])

        for col in reversed(
            [
                "timestamp",
                "owner",
                "name",
                "twitter_id",
                "twitter_name",
                "twitter_username",
                "note",
                "url",
                "war_name",
                "quest_name",
                "runs",
            ]
        ):
            new_row.insert(0, row[col][0])

        new_row.insert(0, idx)
        ws.append(new_row)


def create_list(wb: Workbook, reports_df: pd.DataFrame) -> None:
    """各カテゴリの報告シートを出力する

    Args:
        wb (Workbook): 出力するワークブック
        reports_df (pd.DataFrame): 報告データ
    """
    order = [
        "修練場",
        "フリクエ1部",
        "フリクエ1.5部",
        "フリクエ2部",
        "奏章",
        "冠位戴冠戦",
        "その他クエスト",
        "Error",
    ]

    # データフレームを 'category' でグループ化
    grouped_reports = reports_df.groupby("category")

    # 指定された順序に従って新しいシートを作成
    for category in order:
        if category in grouped_reports.groups:
            category_group_df = grouped_reports.get_group(category)
        else:
            category_group_df = (
                pd.DataFrame()
            )  # データが存在しない場合、空のデータフレームを作成

        ws = wb.create_sheet(title=category)

        append_rows_to_sheet(ws, category_group_df)
